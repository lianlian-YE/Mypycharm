#!/usr/bin/env python
# coding:utf-8


""""
Name: 04_读取excel表的内容.py
Date: 2018/06/03
Connect: xc_guofan@163.com
Author: lvah
Desc:


"""
import os

import openpyxl
def readwb(wbname, sheetname=None):
    wb = openpyxl.load_workbook(wbname)
    if not sheetname:
        sheet = wb.active
    else:
        sheet = wb[sheetname]

    all_info = []
    # 依次遍历工作表的每一行;返回的是元组里面多个Cell对象;
    for row in sheet.rows:
        # 通过列表生成式， 获取每一行的每个单元格的内容;
        child = [cell.value for cell in row]
        all_info.append(child)

    return sorted(all_info, key=lambda item: item[2])



def createwb(wbname):
    # 判断是否有后缀名， 如果没有,则增加;
    if not wbname.endswith('.xlsx'):
        wbname = wbname + '.xlsx'
    wb = openpyxl.Workbook()
    # 保存excel表
    wb.save(filename=wbname)
    print("新建Excel[%s]成功!" %(wbname))

def save_to_excel(data, wbname, sheetname='Sheet1'):
    """
    保存排序好的数据信息到新的excel表中
    """
    createwb(wbname)
    print("写入Excel[%s]中...." %(wbname))
    wb = openpyxl.load_workbook(wbname)
    sheet = wb.active
    # 修改工作表的名称;
    sheet.title = sheetname

    #
    # data = [
    #     ['xxx', 'apple', 12],
    #     ['xxx', 'apple', 12],
    #     ['xxx', 'apple', 12],
    # ]
    #
    # row代表索引, item代表一行内容; eg:['xxx', 'apple', 12];
    for row, item in enumerate(data):
        # column代表索引, cellValue代表每个单元格的内容, eg:'apple'
        for column, cellValue in enumerate(item):
            # 将数据信息写入excel单元格;
            # 单元格的行和列没有0;
            sheet.cell(row=row+1, column=column+1, value=cellValue)

    wb.save(filename=wbname)
    print("保存成功!")

sorted_info = readwb('excelDemo/example.xlsx', 'example')
save_to_excel(sorted_info,'excelDemo/sorted_example.xlsx' )


