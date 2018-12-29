#!/usr/bin/env python
#coding:utf-8


""""
Name: 05_写入数据到excel表中.py
Date: 2018/06/03
Connect: xc_guofan@163.com
Author: lvah
Desc:


"""
import os

import openpyxl


def createwb(wbname):
    try:
        if not wbname.endswith('.xlsx'):
            wbname = wbname + '.xlsx'
        wb = openpyxl.Workbook()
        wb.save(filename=wbname)
        print("新建Excel[%s]成功!" %(wbname))
    except Exception as e:
        print(e)


def save_to_excel(sheetname, wbname):
    print("写入Excel[%s]中...." %(wbname))
    wb = openpyxl.load_workbook(filename=wbname)
    sheet = wb.active
    sheet.title = sheetname


    for row in range(3):
        for column in range(4):
            sheet.cell(row=row+1, column=column+1, value=1)

    wb.save(filename=wbname)
    print("保存成功!")



createwb('hello')
save_to_excel('example', 'hello.xlsx')








