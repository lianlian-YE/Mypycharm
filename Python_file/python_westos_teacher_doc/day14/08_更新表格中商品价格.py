#!/usr/bin/env python
# coding:utf-8


""""
Name: 08_更新表格中商品价格.py
Date: 2018/06/03
Connect: xc_guofan@163.com
Author: lvah
Desc:

每一行代表一次单独的销售。列分别是销售产品的类型(A)、产品每磅的价格
(B)、销售的磅数(C),以及这次销售的总收入。TOTAL 列设置为 Excel 公式,将每磅的成本乘以销售的磅数,并将结果取整到分。有了这个公式,如果列 B 或 C 发生变化,TOTAL 列中的单元格将自动更新.



需要更新的价格如下:
Celery	1.19
Garlic	3.07
Lemon	1.27

现在假设 Garlic、 Celery 和 Lemons 的价格输入的不正确。这让你面对一项无聊
的任务:遍历这个电子表格中的几千行,更新所有 garlic、celery 和 lemon 行中每磅
的价格。你不能简单地对价格查找替换,因为可能有其他的产品价格一样,你不希
望错误地“更正”。对于几千行数据,手工操作可能要几小时。





"""

# 不建议使用;
# productName = ""
#
# if productName == 'Celery':
#     cellObj = 1.19
# elif     productName == "Garlic":
#     cellObj =3.07
# elif productName ==  "Lemon":
#     cellObj =1.27
#


# 需要更新的商品名称和价格组成的字典;
priceUpdate = {
    'Celery': 1.19,
    "Garlic": 3.07,
    "Lemon": 1.27
}

import openpyxl

# 需要操作的excel表;
filename = 'excelDemo/produceSales.xlsx'

# 加载excel表格;
wb  = openpyxl.load_workbook(filename)
# 打开当前的工作表;
sheet = wb.active

# 依次遍历excel的每一行;index代表索引值; row代表每一行内容;
for index, row in enumerate(sheet.rows):
    # 拿出第index+1行， 第一列的单元格内容, 获取指定行行的商品名称;
    produceName = sheet.cell(row=index+1, column=1).value
    # 判断商品名称是否在需要更新的字典里?
    if produceName in priceUpdate:
        # 如果需要更新， 从字典中获取更新的商品价格;
        updatePrice = priceUpdate[produceName]
        # 显示信息；
        print("更新 %s 行%s的价格为%s" %(index+1, produceName, updatePrice))
        # 更新指定行指定列商品的价格;
        sheet.cell(row=index+1, column=2).value = updatePrice
# 保存excel表;
wb.save(filename=filename)