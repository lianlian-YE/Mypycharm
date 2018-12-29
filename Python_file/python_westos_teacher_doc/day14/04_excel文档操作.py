#!/usr/bin/env python
#coding:utf-8


""""
Name: 04_excel文档操作.py
Date: 2018/06/03
Connect: xc_guofan@163.com
Author: lvah
Desc:


**********************基本概念*************************
    # 工作薄:workbook

    # 工作表:sheet

    # 活动表: 打开工作薄默认所在的工作表；


    # 列(column)： A B C D E
    # 行(row)： 1 2 3 4 5 6
    # 单元格(cell)




*****************基本操作*******************************

# 导入工作薄
1. openpyxl.load_workbook('excelDemo/example.xlsx')
2. wb.sheetnames
3. wb.active
4. wb['Sheet1']

# 工作表
5. sheet.title
sheet.cell(row=1, column=2)

# 单元格
6. cell = sheet['A1']
7. cell_value = sheet['A1'].value
8. cell.row, cell.column cell.coordinate







"""

import openpyxl


# 1. 读取excel文档
wb = openpyxl.load_workbook('excelDemo/example.xlsx')

# # 返回一个workbook对象， 有点类似于文件对象;
# print(wb, type(wb))



# 2. 在工作薄中取得工作表
# print(wb.get_sheet_names())
# 返回一个列表， 存储excel表中所有的sheet工作表;
print(wb.sheetnames)

# 返回一个worksheet对象， 返回当前的活动表;
# print(wb.get_active_sheet())
# print(wb.active)



# 3. 获取工作表中， 单元格的信息
# wb.get_sheet_by_name('Sheet1')
sheet = wb['example']
print(sheet['A1'])
print(sheet['B1'].value)

cell = sheet['B1']
print(cell.row, cell.column)


print(sheet.cell(row=3, column=2))
print(sheet.cell(row=3, column=2).value)
print(sheet.cell(row=3, column=2, value='www'))


# sheet的属性

print(sheet.max_column)
print(sheet.max_row)
print(sheet.title)
sheet.title = 'example'
print(sheet.title)


for row in sheet.rows:
    for cell in row:
        print(cell.value, end='\t')
    print('\n')

wb.save(filename="excelDemo/example.xlsx")





